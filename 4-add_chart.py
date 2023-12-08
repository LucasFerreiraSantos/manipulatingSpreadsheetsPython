from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

#1 - Read workbook and spreadsheet
wb = load_workbook("./data/Pivot_table.xlsx")
sheet = wb['Relatório']

#2 - Row and column references
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row
print(min_column, max_column)
print(min_row, max_row)

# Add data and category in grafic
barchart = BarChart()

data = Reference(
  sheet,
  min_col=min_column + 1,
  max_col=max_column,
  min_row=min_row,
  max_row=max_row
)

categories = Reference(
  sheet,
  min_col=min_column,
  max_col=min_column,
  min_row=min_row + 1,
  max_row=max_row
)

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

#4 - Create the grafic
sheet.add_chart(barchart, "B10")
barchart.title = "Vendas por Fabricantes"
barchart.style = 2

#Save WorkBook
wb.save("./data/barchart.xlsx")